#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl as px
import stm as engine

W = px.load_workbook('English_to_Bangla_Parallel_Corpus.xlsx')
p = W.get_sheet_by_name(name = 'Sheet1')

a=[]
lan1 = ''
lan2 = ''
val1 = ''
val2 = ''
j = 0
f = open('myfile.txt', 'w')
f.write("[\n")
mxr =  p.max_row
d = 0
for row in p.iter_rows():
    i = 0
    if j == 1:
    	f.write("  {\n")
    for k in row:
        if i == 0:
            if j == 0:
                lan1 = k.internal_value.encode('utf-8')
            else:
                val1 = k.internal_value.encode('utf-8')
                val1.replace('"','\\"')
                f.write("    \"" + lan1 + "\": " + "\"" + val1 + "\",\n")
        else:
            if j == 0:
                lan2 = k.internal_value.encode('utf-8')
                j = 1
            else:
                val2 = k.internal_value.encode('utf-8')
                #val2 = engine.convert(val2).encode('utf-8')
                print val2
                val2.replace('"','\\"')
                f.write("    \"" + lan2 + "\": " + "\"" + val2 + "\"\n")
                if d == mxr - 1:
                	f.write("  }\n")
                else:
                	f.write("  },\n")
        i = i + 1
    d = d + 1       
f.write("]")
f.close()









